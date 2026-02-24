import csv
import io
import json
from datetime import datetime

from models import WalletReport


def to_csv(report: WalletReport) -> bytes:
    """Export wallet report to CSV."""
    out = io.StringIO()
    w = csv.writer(out)

    w.writerow(["WEB3 WALLET ANALYSIS REPORT"])
    w.writerow(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    w.writerow([])

    # ── Summary ───────────────────────────────────────────────────────
    w.writerow(["SUMMARY"])
    w.writerow(["Address", report.address])
    w.writerow(["Address Type", report.address_type.upper()])
    w.writerow(["Total Transactions", report.total_transactions])
    w.writerow(["Total Received (USD)", f"${report.total_received_usd:,.2f}"])
    w.writerow(["Total Sent (USD)", f"${report.total_sent_usd:,.2f}"])
    w.writerow(["Net Flow (USD)", f"${report.net_flow_usd:,.2f}"])
    w.writerow(["Total Gas Spent (USD)", f"${report.total_gas_spent_usd:,.2f}"])
    w.writerow(["Wallet Age (Days)", report.wallet_age_days or "N/A"])
    w.writerow(["First Activity", str(report.first_activity or "N/A")])
    w.writerow(["Last Activity", str(report.last_activity or "N/A")])
    w.writerow(["Active Chains", ", ".join(report.chains_with_activity)])
    w.writerow([])

    # ── Chain Breakdown ───────────────────────────────────────────────
    w.writerow(["CHAIN BREAKDOWN (Sorted by Activity)"])
    w.writerow([
        "Chain", "Symbol", "Transactions", "Incoming", "Outgoing",
        "Received (Native)", "Received (USD)", "Sent (Native)", "Sent (USD)",
        "Gas (Native)", "Gas (USD)", "Token Transfers", "Contracts Interacted",
        "First Tx", "Last Tx",
    ])

    for s in report.chain_summaries:
        w.writerow([
            s.chain_name, s.native_symbol, s.total_transactions,
            s.incoming_transactions, s.outgoing_transactions,
            f"{s.total_received:.6f}", f"${s.total_received_usd:,.2f}",
            f"{s.total_sent:.6f}", f"${s.total_sent_usd:,.2f}",
            f"{s.total_gas_spent:.6f}", f"${s.total_gas_spent_usd:,.2f}",
            s.token_transfers_count, s.unique_contracts_interacted,
            str(s.first_transaction_date or "N/A"),
            str(s.last_transaction_date or "N/A"),
        ])

    w.writerow([])

    # ── AI Insights ───────────────────────────────────────────────────
    if report.ai_insights:
        w.writerow(["AI INSIGHTS"])
        for line in report.ai_insights.split("\n"):
            w.writerow([line])

    return out.getvalue().encode("utf-8")


def to_json(report: WalletReport) -> bytes:
    """Export wallet report as formatted JSON."""
    return json.dumps(report.model_dump(), indent=2, default=str).encode("utf-8")


def to_excel(report: WalletReport) -> bytes:
    """Export wallet report to formatted Excel workbook."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    # ── Summary Sheet ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    accent = PatternFill(start_color="6c5ce7", end_color="6c5ce7", fill_type="solid")
    dark = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)

    ws.merge_cells("A1:E1")
    ws["A1"] = "Web3 Wallet Analysis Report"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = accent
    ws["A1"].alignment = Alignment(horizontal="center")

    rows = [
        ("Address", report.address),
        ("Type", report.address_type.upper()),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("Total Transactions", f"{report.total_transactions:,}"),
        ("Total Received (USD)", f"${report.total_received_usd:,.2f}"),
        ("Total Sent (USD)", f"${report.total_sent_usd:,.2f}"),
        ("Net Flow (USD)", f"${report.net_flow_usd:,.2f}"),
        ("Total Gas Spent (USD)", f"${report.total_gas_spent_usd:,.2f}"),
        ("Active Chains", str(len(report.chains_with_activity))),
        ("Wallet Age (Days)", str(report.wallet_age_days or "N/A")),
        ("First Activity", str(report.first_activity or "N/A")),
        ("Last Activity", str(report.last_activity or "N/A")),
    ]
    for i, (label, value) in enumerate(rows, 3):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = bold
        ws[f"B{i}"] = value

    # ── Chain Breakdown Sheet ─────────────────────────────────────────
    ws2 = wb.create_sheet("Chain Breakdown")
    headers = [
        "Chain", "Symbol", "Transactions", "Received (USD)", "Sent (USD)",
        "Gas (USD)", "Token Transfers", "Contracts", "First Activity", "Last Activity",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = white_bold
        cell.fill = dark

    for i, s in enumerate(report.chain_summaries, 2):
        ws2.cell(row=i, column=1, value=s.chain_name)
        ws2.cell(row=i, column=2, value=s.native_symbol)
        ws2.cell(row=i, column=3, value=s.total_transactions)
        ws2.cell(row=i, column=4, value=f"${s.total_received_usd:,.2f}")
        ws2.cell(row=i, column=5, value=f"${s.total_sent_usd:,.2f}")
        ws2.cell(row=i, column=6, value=f"${s.total_gas_spent_usd:,.2f}")
        ws2.cell(row=i, column=7, value=s.token_transfers_count)
        ws2.cell(row=i, column=8, value=s.unique_contracts_interacted)
        ws2.cell(row=i, column=9, value=str(s.first_transaction_date or "N/A"))
        ws2.cell(row=i, column=10, value=str(s.last_transaction_date or "N/A"))

    # Auto-fit column widths
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
